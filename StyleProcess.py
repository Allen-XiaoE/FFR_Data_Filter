import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, colors, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import IconSet, FormatObject, Rule
import warnings
warnings.filterwarnings('ignore')

def styleprocess(robottype,path):
    os.chdir(path)
    #1.读取Excel文件
    file = r'{0}_DataResource.xlsx'.format(robottype)
    wb=openpyxl.load_workbook(file)
    Failure_Data = wb.get_sheet_by_name('Failure Data')
    Delivery_Data = wb.get_sheet_by_name('Delivery Data')
    DMAIC_Data = wb.get_sheet_by_name('DMAIC Data')
    Failure_Category = wb.get_sheet_by_name('Failure Category')

    #2.加边框
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                            right=Side(style='thin', color=colors.BLACK),
                            top=Side(style='thin', color=colors.BLACK),
                            bottom=Side(style='thin', color=colors.BLACK))

    fill_blue = PatternFill("solid", fgColor="3399FF")#yellow     #对不同区块分开进行颜色的设定
    fill_green = PatternFill("solid", fgColor="66FF66")#green     #第一个双引号内为颜色填充类型，第二个则为RGB颜色        
    fill_lightblue = PatternFill("solid", fgColor="66FFFF")
    fill_yellow = PatternFill("solid", fgColor="FFFF66")

    fontstyle = Font(name = 'Calibri',size=11)

    #Failure_Data表格设置
    #1.1设置Border和字体
    for row in Failure_Data.rows:
        for c in row:
            c.border = border_set
            c.font = fontstyle
            c.alignment = Alignment(horizontal='center',vertical='center')
            

    #1.2 设置列宽
    for i in range(Failure_Data.max_column):
        colname = get_column_letter(i + 1)
        if i + 1 <= 5:
            Failure_Data.column_dimensions[colname].width=18
        elif i + 1 <= 8:
            Failure_Data.column_dimensions[colname].width=22
        elif i + 1 in [9,11,18,23]:
            Failure_Data.column_dimensions[colname].width=60
            
        elif i + 1 in [10,14,16]:
            Failure_Data.column_dimensions[colname].width=18
        elif i + 1 == 26:
            Failure_Data.column_dimensions[colname].width=5
        else:
            Failure_Data.column_dimensions[colname].width=22

    #1.3 设置行高
    for i in range(Failure_Data.max_row):
        row = Failure_Data.row_dimensions[i+1]
        if i + 1 == 1:
            row.height = 40
        else:
            row.height = 18

    #1.4 设置颜色
    failurecell_blue = Failure_Data['A1:Y1']
    for row in failurecell_blue:
        for c in row:
            c.fill = fill_blue

    failurecell_green = Failure_Data['A2:R'+str(Failure_Data.max_row)]
    for row in failurecell_green:
        for c in row:
            c.fill = fill_green

    failurecell_yellow = Failure_Data['S2:V'+str(Failure_Data.max_row)]
    for row in failurecell_yellow:
        for c in row:
            c.fill = fill_yellow

    failurecell_lighblue= Failure_Data['W2:Y'+str(Failure_Data.max_row)]
    for row in failurecell_lighblue:
        for c in row:
            c.fill = fill_lightblue

    #1.5 设置对齐方式
    left_alignment_cells = Failure_Data['G2:W{0}'.format(Failure_Data.max_row)]
    for row in left_alignment_cells:
        for c in row:
            c.alignment = Alignment(horizontal='left',vertical='center')
    

    #1.5 写公式
    for i in range(1,Failure_Data.max_row):
        col = i + 1
        failure_formula=r'=IF(LEFT(SUBSTITUTE(SUBSTITUTE(IF(ISBLANK(S{0}),"_$",S{0})&"  "&IF(ISBLANK(T{0}),"$",T{0})&"  "&IF(ISBLANK(U{0}),"$",U{0})&"  "&IF(ISBLANK(V{0}),"$",V{0}),"  ","_"),"_$",""),1)="_",RIGHT(SUBSTITUTE(SUBSTITUTE(IF(ISBLANK(S{0}),"_$",S{0})&"  "&IF(ISBLANK(T{0}),"$",T{0})&"  "&IF(ISBLANK(U{0}),"$",U{0})&"  "&IF(ISBLANK(V{0}),"$",V{0}),"  ","_"),"_$",""),LEN(SUBSTITUTE(SUBSTITUTE(IF(ISBLANK(S{0}),"_$",S{0})&"  "&IF(ISBLANK(T{0}),"$",T{0})&"  "&IF(ISBLANK(U{0}),"$",U{0})&"  "&IF(ISBLANK(V{0}),"$",V{0}),"  ","_"),"_$",""))-1),SUBSTITUTE(SUBSTITUTE(IF(ISBLANK(S{0}),"_$",S{0})&"  "&IF(ISBLANK(T{0}),"$",T{0})&"  "&IF(ISBLANK(U{0}),"$",U{0})&"  "&IF(ISBLANK(V{0}),"$",V{0}),"  ","_"),"_$",""))'.format(col)
        bulb_formula=r'= IF((W{0}<>""),1,0)'.format(col)
        Failure_Data['W' + str(col)] = failure_formula
        Failure_Data['Z' + str(col)] = bulb_formula
    Failure_Data['Z1'] = r'=IF(COUNTIF(Z2:Z{0},0) = 0,1,0)'.format(Failure_Data.max_row)

    Failure_Data['AA1'] = 1
    Failure_Data.column_dimensions['AA'].hidden = True
    #1.6 设置三色灯
    first = FormatObject(type='percent', val=0)
    second = FormatObject(type='percent', val=33)
    third = FormatObject(type='percent', val=67)
    iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third], showValue=False, percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    formularange='Z1:Z'+str(Failure_Data.max_row)
    Failure_Data.conditional_formatting.add(formularange, rule)

    #2.Delivery_Data表格设置
    #2.1 设置Border
    for row in Delivery_Data.rows:
        for c in row:
            c.border = border_set
            c.alignment = Alignment(horizontal='center',vertical='center')
            c.fill = fill_green
            c.font = fontstyle

    #2.2 设置填充色
    deliverycell = Delivery_Data['A1:I1']
    for row in deliverycell:
        for c in row:
            c.fill = fill_blue

    #2.3 设置列宽  
    for i in range(Delivery_Data.max_column):
        colname = get_column_letter(i + 1)
        Delivery_Data.column_dimensions[colname].width=22


    #3.DMAIC_Data表格
    #3.1 设置Border
    for row in DMAIC_Data.rows:
        for c in row:
            c.border = border_set
            c.font = fontstyle
            c.fill = fill_green
            c.alignment = Alignment(horizontal='center',vertical='center')

    #3.2 设置填充颜色
    dmaiccell = DMAIC_Data['A1:J1']
    for row in dmaiccell:
        for c in row:
            c.fill = fill_blue

    #3.3 设置列宽
    for i in range(DMAIC_Data.max_column):
        colname = get_column_letter(i + 1)
        if i + 1 == 2:
            DMAIC_Data.column_dimensions[colname].width=30
        else:
            DMAIC_Data.column_dimensions[colname].width=10
    DMAIC_Data['A1'] = 'Rank'

    #Failure Catogery排序，并设置最新DataValidation
    validationlist = []
    for col in range(Failure_Category.max_column):

        colname = get_column_letter(col + 1)
        list = []
        for i in Failure_Category[colname][1:]:
            if i.value != None:
                list.append(i.value)
        list.sort(key=None,reverse=False)
        listlength = len(list)
        dv = DataValidation(type="list", formula1="'Failure Category'!${0}$2:${0}${1}".format(colname,listlength+1), allow_blank=True)#F!$A$1:$A$26
        validationlist.append(dv)

        for j in range(listlength):
            Failure_Category[colname + str(j + 2)] = list[j]
        
    for i in range(len(validationlist)):
        validationcolname = get_column_letter(i+19)
        dv = validationlist[i]
        Failure_Data.add_data_validation(dv)
        for col in range(1,Failure_Data.max_row):
            dv.add(validationcolname+str(col+1))

    Failure_Category.sheet_state = "hidden"
    if DMAIC_Data.max_row > 1:
        dv1 = DataValidation(type='list',formula1='"D, M, A, I, C"',allow_blank=True)
        DMAIC_Data.add_data_validation(dv1)
        for col in range(1,DMAIC_Data.max_row):
            dv1.add('D'+str(col+1))

    wb.save(file)

    print('{0} StyleProcess is Done!'.format(robottype))
    print('---------------------------------\n')